import os
import sys
import requests
import argparse
from bs4 import BeautifulSoup
import google.generativeai as genai
from dotenv import load_dotenv
from fpdf import FPDF
from docx import Document
from openpyxl import Workbook
from PIL import Image, ImageDraw, ImageFont

def configurar_api():
    """Carrega a chave da API do arquivo .env e configura o SDK do Gemini."""
    load_dotenv()
    api_key = os.getenv("GOOGLE_API_KEY")
    if not api_key:
        raise ValueError("A chave da API do Google não foi encontrada. Verifique seu arquivo .env.")
    genai.configure(api_key=api_key)
    return genai.GenerativeModel('gemini-1.5-flash')

def extrair_texto_do_artigo(url):
    """
    Acessa uma URL, baixa o conteúdo HTML e extrai o texto dos parágrafos.
    """
    print(f"Acessando o artigo em: {url}")
    try:
        response = requests.get(url, headers={'User-Agent': 'Mozilla/5.0'})
        response.raise_for_status()
        soup = BeautifulSoup(response.content, 'html.parser')
        paragrafos = soup.find_all('p')
        texto_artigo = '\n'.join([p.get_text() for p in paragrafos])
        if not texto_artigo:
            print("Aviso: Nenhum texto de parágrafo (<p>) encontrado na página.")
            return soup.get_text(separator='\n', strip=True)
        return texto_artigo
    except requests.exceptions.RequestException as e:
        print(f"Erro ao acessar a URL: {e}")
        return None

def resumir_texto(modelo, texto, idioma, estilo):
    """
    Envia o texto para o modelo Gemini e retorna o resumo no idioma e estilo especificados.
    """
    if not texto:
        return "Não foi possível extrair texto para resumir."
    
    print(f"Enviando texto para a IA para sumarização em {idioma} no estilo '{estilo}'... Isso pode levar um momento.")
    
    prompt = f"Por favor, resuma o seguinte artigo em {idioma}. O estilo do resumo deve ser: '{estilo}'.\n\n-- INÍCIO DO ARTIGO --\n{texto}\n-- FIM DO ARTIGO --"
    
    try:
        response = modelo.generate_content(prompt)
        return response.text
    except Exception as e:
        return f"Ocorreu um erro durante a sumarização: {e}"

def salvar_como_txt(texto, nome_arquivo):
    """Salva o texto fornecido em um arquivo .txt."""
    try:
        with open(nome_arquivo, 'w', encoding='utf-8') as f:
            f.write(texto)
        print(f"Resumo salvo com sucesso em: {nome_arquivo}")
    except IOError as e:
        print(f"Erro ao salvar o arquivo .txt: {e}")

def salvar_como_pdf(texto, nome_arquivo):
    """Salva o texto fornecido em um arquivo .pdf."""
    try:
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Helvetica", size=12)
        pdf.multi_cell(0, 10, text=texto.encode('latin-1', 'replace').decode('latin-1'))
        pdf.output(nome_arquivo)
        print(f"Resumo salvo com sucesso em: {nome_arquivo}")
    except Exception as e:
        print(f"Erro ao salvar o arquivo .pdf: {e}")

def salvar_como_docx(texto, nome_arquivo):
    """Salva o texto fornecido em um arquivo .docx."""
    try:
        document = Document()
        document.add_paragraph(texto)
        document.save(nome_arquivo)
        print(f"Resumo salvo com sucesso em: {nome_arquivo}")
    except Exception as e:
        print(f"Erro ao salvar o arquivo .docx: {e}")

def salvar_como_xlsx(texto, nome_arquivo):
    """Salva o texto fornecido em um arquivo .xlsx."""
    try:
        wb = Workbook()
        ws = wb.active
        ws['A1'] = "Resumo do Artigo"
        ws['A2'] = texto
        wb.save(nome_arquivo)
        print(f"Resumo salvo com sucesso em: {nome_arquivo}")
    except Exception as e:
        print(f"Erro ao salvar o arquivo .xlsx: {e}")

def salvar_como_imagem(texto, nome_arquivo, tipo_imagem='png'):
    """Salva o texto fornecido em um arquivo de imagem (PNG ou JPG)."""
    try:
        # Configurações da imagem
        largura = 800
        padding = 50
        tamanho_fonte = 20
        cor_fundo = (255, 255, 255)  # Branco
        cor_texto = (0, 0, 0)      # Preto

        # Tenta carregar uma fonte padrão do sistema
        try:
            fonte = ImageFont.truetype("arial.ttf", tamanho_fonte)
        except IOError:
            # Fallback para uma fonte genérica se arial.ttf não for encontrada
            fonte = ImageFont.load_default()

        # Calcula a altura necessária para o texto
        linhas = []
        draw_temp = ImageDraw.Draw(Image.new('RGB', (1, 1)))
        
        # Quebra o texto em linhas que caibam na largura da imagem
        palavras = texto.split(' ')
        linha_atual = []
        for palavra in palavras:
            linha_tentativa = " ".join(linha_atual + [palavra])
            largura_texto, altura_texto = draw_temp.textbbox((0,0), linha_tentativa, font=fonte)[2:]
            if largura_texto <= (largura - 2 * padding):
                linha_atual.append(palavra)
            else:
                linhas.append(" ".join(linha_atual))
                linha_atual = [palavra]
        linhas.append(" ".join(linha_atual))

        altura_total_texto = len(linhas) * (tamanho_fonte + 5) # +5 para espaçamento entre linhas
        altura = altura_total_texto + 2 * padding
        if altura < 200: # Altura mínima para imagens muito curtas
            altura = 200

        img = Image.new('RGB', (largura, altura), color = cor_fundo)
        d = ImageDraw.Draw(img)

        y_texto = padding
        for linha in linhas:
            d.text((padding, y_texto), linha, font=fonte, fill=cor_texto)
            y_texto += (tamanho_fonte + 5)

        img.save(nome_arquivo, tipo_imagem.upper())
        print(f"Resumo salvo com sucesso em: {nome_arquivo}")
    except Exception as e:
        print(f"Erro ao salvar o arquivo de imagem: {e}")

def main():
    """
    Função principal que orquestra o processo de sumarização e salvamento.
    """
    parser = argparse.ArgumentParser(description="Sumarizador de Artigos com IA usando Google Gemini.")
    parser.add_argument("url", help="A URL do artigo a ser resumido.")
    parser.add_argument("--language", default="português", help="O idioma do resumo (ex: 'inglês', 'espanhol'). Padrão: português.")
    parser.add_argument("--style", default="um parágrafo conciso", help="O estilo do resumo (ex: 'em tópicos', 'para um executivo').")
    parser.add_argument("--output", help="O nome do arquivo para salvar o resumo (ex: resumo.txt, resumo.pdf, resumo.docx, resumo.xlsx, resumo.png, resumo.jpg).")
    
    args = parser.parse_args()
    
    try:
        modelo = configurar_api()
        texto_artigo = extrair_texto_do_artigo(args.url)
        
        if texto_artigo:
            resumo = resumir_texto(modelo, texto_artigo, args.language, args.style)
            
            print(f"\n--- Resumo do Artigo (Estilo: {args.style}) ---")
            print(resumo)
            print("----------------------------------------------------\n")
            
            if args.output:
                nome_arquivo = args.output.lower()
                if nome_arquivo.endswith('.txt'):
                    salvar_como_txt(resumo, nome_arquivo)
                elif nome_arquivo.endswith('.pdf'):
                    salvar_como_pdf(resumo, nome_arquivo)
                elif nome_arquivo.endswith('.docx'):
                    salvar_como_docx(resumo, nome_arquivo)
                elif nome_arquivo.endswith('.xlsx'):
                    salvar_como_xlsx(resumo, nome_arquivo)
                elif nome_arquivo.endswith('.png'):
                    salvar_como_imagem(resumo, nome_arquivo, 'png')
                elif nome_arquivo.endswith('.jpg') or nome_arquivo.endswith('.jpeg'):
                    salvar_como_imagem(resumo, nome_arquivo, 'jpeg')
                else:
                    print("Formato de arquivo não suportado. Use .txt, .pdf, .docx, .xlsx, .png ou .jpg.")
            
    except ValueError as e:
        print(f"Erro de configuração: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()