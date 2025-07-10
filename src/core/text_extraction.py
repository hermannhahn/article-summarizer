"""Extracts text from various sources like URLs and local files.

This module provides functionality to extract plain text from web articles,
PDF documents, DOCX files, and XLSX spreadsheets. It handles network requests
for URLs and uses appropriate libraries for local file parsing.
"""

import logging
import os

import requests
from bs4 import BeautifulSoup
from docx import Document
from openpyxl import load_workbook
from PyPDF2 import PdfReader

# --- Private Helper Functions ---


def _extract_text_from_pdf(file_path: str) -> str | None:
    """Extract text from a PDF file.

    Args:
        file_path (str): The absolute path to the PDF file.

    Returns:
        str | None: The extracted text as a single string, or None if extraction fails.
    """
    try:
        reader = PdfReader(file_path)
        text = "".join(page.extract_text() + "\n" for page in reader.pages)
        logging.info(f"Successfully extracted text from PDF: {file_path}")
        return text
    except FileNotFoundError:
        logging.error(f"PDF file not found at: {file_path}")
        return None
    except Exception as e:
        logging.error(f"Failed to extract text from PDF '{file_path}': {e}")
        return None


def _extract_text_from_docx(file_path: str) -> str | None:
    """Extract text from a DOCX file.

    Args:
        file_path (str): The absolute path to the DOCX file.

    Returns:
        str | None: The extracted text as a single string, or None if extraction fails.
    """
    try:
        document = Document(file_path)
        text = "\n".join([paragraph.text for paragraph in document.paragraphs])
        logging.info(f"Successfully extracted text from DOCX: {file_path}")
        return text
    except FileNotFoundError:
        logging.error(f"DOCX file not found at: {file_path}")
        return None
    except Exception as e:
        logging.error(f"Failed to extract text from DOCX '{file_path}': {e}")
        return None


def _extract_text_from_xlsx(file_path: str) -> str | None:
    """Extract text from all cells in an XLSX file.

    Args:
        file_path (str): The absolute path to the XLSX file.

    Returns:
        str | None: The concatenated text from all cells, or None if extraction fails.
    """
    try:
        workbook = load_workbook(file_path, read_only=True)
        text_parts = []
        for sheet in workbook:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        text_parts.append(str(cell.value))
        logging.info(f"Successfully extracted text from XLSX: {file_path}")
        return " ".join(text_parts)
    except FileNotFoundError:
        logging.error(f"XLSX file not found at: {file_path}")
        return None
    except Exception as e:
        logging.error(f"Failed to extract text from XLSX '{file_path}': {e}")
        return None


# --- Public API ---


def extract_text(source: str) -> str | None:
    """Extract text from a given source (URL or local file path).

    This function determines the type of source and delegates to the
    appropriate helper function for text extraction.

    Args:
        source (str): The URL or the local file path (PDF, DOCX, XLSX).

    Returns:
        str | None: The extracted text, or None if the source is invalid or
                    text extraction fails.
    """
    if source.startswith(("http://", "https://")):
        logging.info(f"Extracting text from URL: {source}")
        try:
            headers = {"User-Agent": "Mozilla/5.0"}
            response = requests.get(source, headers=headers, timeout=15)
            response.raise_for_status()
            soup = BeautifulSoup(response.content, "html.parser")
            paragraphs = soup.find_all("p")
            article_text = "\n".join(p.get_text() for p in paragraphs)
            if not article_text.strip():
                logging.warning(
                    "No <p> tags found. Falling back to generic text extraction."
                )
                return soup.get_text(separator="\n", strip=True)
            return article_text
        except requests.exceptions.HTTPError as e:
            logging.error(f"HTTP Error accessing URL '{source}': {e}")
            return None
        except requests.exceptions.ConnectionError as e:
            logging.error(f"Connection Error for URL '{source}': {e}")
            return None
        except requests.exceptions.Timeout as e:
            logging.error(f"Request timed out for URL '{source}': {e}")
            return None
        except requests.exceptions.RequestException as e:
            logging.error(
                f"An unexpected network error occurred for URL '{source}': {e}"
            )
            return None

    elif os.path.exists(source):
        logging.info(f"Extracting text from local file: {source}")
        file_extension = os.path.splitext(source)[1].lower()
        if file_extension == ".pdf":
            return _extract_text_from_pdf(source)
        elif file_extension == ".docx":
            return _extract_text_from_docx(source)
        elif file_extension == ".xlsx":
            return _extract_text_from_xlsx(source)
        else:
            logging.warning(
                f"Unsupported local file type: '{file_extension}'. Only .pdf, .docx, and .xlsx are supported."
            )
            return None
    else:
        logging.error(
            f"Source not found: '{source}'. It is not a valid URL or an existing file path."
        )
        return None
