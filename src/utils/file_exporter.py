"""
Handles the export of summaries to various file formats.

This module provides functions to save text content into different file types
such as TXT, PDF, DOCX, XLSX, and image formats (PNG, JPG). It includes
error handling for file operations and ensures that the target directory
for the output file exists before writing.
"""

import os
import logging
from fpdf import FPDF
from docx import Document
from openpyxl import Workbook
from PIL import Image, ImageDraw, ImageFont

def _ensure_directory_exists(filename):
    """
    Ensures that the directory for the given filename exists.

    If the directory does not exist, it will be created.

    Args:
        filename (str): The path to the file.
    """
    directory = os.path.dirname(filename)
    if directory:
        os.makedirs(directory, exist_ok=True)

def save_as_txt(text, filename):
    """
    Saves the provided text to a .txt file.

    Args:
        text (str): The text content to save.
        filename (str): The path to the output .txt file.
    """
    try:
        _ensure_directory_exists(filename)
        with open(filename, 'w', encoding='utf-8') as f:
            f.write(text)
        logging.info(f"Summary successfully saved to: {filename}")
    except IOError as e:
        logging.error(f"Failed to save .txt file '{filename}': {e}")
    except Exception as e:
        logging.error(f"An unexpected error occurred while saving .txt file '{filename}': {e}")

def save_as_pdf(text, filename):
    """
    Saves the provided text to a .pdf file.

    Note:
        fpdf2 may have issues with non-latin characters without proper font setup.
        This implementation uses a basic latin-1 encoding with replacement for
        unsupported characters.

    Args:
        text (str): The text content to save.
        filename (str): The path to the output .pdf file.
    """
    try:
        _ensure_directory_exists(filename)
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Helvetica", size=12)
        # Encode to latin-1, replacing unsupported characters.
        encoded_text = text.encode('latin-1', 'replace').decode('latin-1')
        pdf.multi_cell(0, 10, text=encoded_text)
        pdf.output(filename)
        logging.info(f"Summary successfully saved to: {filename}")
    except Exception as e:
        logging.error(f"An unexpected error occurred while saving .pdf file '{filename}': {e}")

def save_as_docx(text, filename):
    """
    Saves the provided text to a .docx file.

    Args:
        text (str): The text content to save.
        filename (str): The path to the output .docx file.
    """
    try:
        _ensure_directory_exists(filename)
        document = Document()
        document.add_paragraph(text)
        document.save(filename)
        logging.info(f"Summary successfully saved to: {filename}")
    except Exception as e:
        logging.error(f"An unexpected error occurred while saving .docx file '{filename}': {e}")

def save_as_xlsx(text, filename):
    """
    Saves the provided text to an .xlsx file.

    The summary is placed in cell A2, with a header in A1.

    Args:
        text (str): The text content to save.
        filename (str): The path to the output .xlsx file.
    """
    try:
        _ensure_directory_exists(filename)
        wb = Workbook()
        ws = wb.active
        ws.title = "Summary"
        ws['A1'] = "Article Summary"
        ws['A2'] = text
        wb.save(filename)
        logging.info(f"Summary successfully saved to: {filename}")
    except Exception as e:
        logging.error(f"An unexpected error occurred while saving .xlsx file '{filename}': {e}")

def save_as_image(text, filename, image_type='png'):
    """
    Saves the provided text to an image file (PNG or JPG).

    This function wraps text to fit within the image width and uses a
    default font if a system-specific one (Arial) is not found.

    Args:
        text (str): The text content to render.
        filename (str): The path to the output image file.
        image_type (str): The format of the image ('png' or 'jpeg').
    """
    try:
        _ensure_directory_exists(filename)
        width, padding, font_size = 800, 50, 20
        bg_color, text_color = (255, 255, 255), (0, 0, 0)

        try:
            font = ImageFont.truetype("arial.ttf", font_size)
        except IOError:
            logging.warning("arial.ttf not found. Falling back to default font for image export.")
            font = ImageFont.load_default()

        # Simple text wrapping
        lines = []
        words = text.split()
        if not words:
            logging.warning("No text provided to save_as_image.")
            return
            
        current_line = words[0]
        for word in words[1:]:
            # Create a temporary ImageDraw to measure text size
            draw_temp = ImageDraw.Draw(Image.new('RGB', (1, 1)))
            bbox = draw_temp.textbbox((0,0), current_line + ' ' + word, font=font)
            if bbox[2] < width - 2 * padding:
                current_line += ' ' + word
            else:
                lines.append(current_line)
                current_line = word
        lines.append(current_line)

        line_height = font.getbbox('A')[3] - font.getbbox('A')[1]
        total_text_height = len(lines) * (line_height + 5)
        height = total_text_height + 2 * padding
        
        img = Image.new('RGB', (width, int(height)), color=bg_color)
        d = ImageDraw.Draw(img)

        y_text = padding
        for line in lines:
            d.text((padding, y_text), line, font=font, fill=text_color)
            y_text += line_height + 5

        img.save(filename, image_type.upper())
        logging.info(f"Summary successfully saved to: {filename}")
    except ImportError:
        logging.error("Pillow library is required for image export. Please install it with 'pip install Pillow'.")
    except Exception as e:
        logging.error(f"An unexpected error occurred while saving image file '{filename}': {e}")
